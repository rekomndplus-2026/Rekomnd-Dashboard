"""
Group Monitor Router
Handles WhatsApp group monitoring, lead detection, and data export.

Endpoints:
  POST /api/monitor/webhook              - Evolution API webhook receiver
  GET  /api/monitor/groups               - List available WhatsApp groups
  GET  /api/monitor/groups/{id}/members  - List group participants
  GET  /api/monitor/groups/{id}/members/export - Export group numbers as Excel
  POST /api/monitor/subscribe            - Register groups + webhook
  GET  /api/monitor/leads                - Get all detected leads (paginated)
  DELETE /api/monitor/leads/{lead_id}    - Delete a lead
  GET  /api/monitor/leads/export         - Export leads as Excel
  GET  /api/monitor/status               - Get monitoring status
"""

import os
import io
import json
import logging
import uuid
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, HTTPException, Request, Response
from fastapi.responses import StreamingResponse

from models.schemas import (
    GroupInfo,
    GroupMember,
    LeadRecord,
    MonitorSubscription,
    WebhookPayload,
)
from services.lead_classifier import classify_message

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/monitor", tags=["Group Monitor"])

LEADS_FILE = "leads.json"

class MockRedis:
    def _load(self):
        if os.path.exists(LEADS_FILE):
            with open(LEADS_FILE, "r") as f:
                try: return json.load(f)
                except: return {}
        return {}
    def _save(self, d):
        with open(LEADS_FILE, "w") as f:
            json.dump(d, f)
    def smembers(self, key):
        return set(self._load().get(key, []))
    def get(self, key):
        return self._load().get(key)
    def setex(self, key, ttl, val):
        d = self._load()
        d[key] = val
        self._save(d)
    def delete(self, key):
        d = self._load()
        if key in d:
            del d[key]
            self._save(d)
    def sadd(self, key, *vals):
        d = self._load()
        s = set(d.get(key, []))
        s.update(vals)
        d[key] = list(s)
        self._save(d)
    def hset(self, key, f, v):
        d = self._load()
        if key not in d: d[key] = {}
        d[key][f] = v
        self._save(d)
    def hgetall(self, key):
        return self._load().get(key, {})
    def hget(self, key, f):
        return self._load().get(key, {}).get(f)
    def hdel(self, key, f):
        d = self._load()
        if key in d and f in d[key]:
            del d[key][f]
            self._save(d)
    def hexists(self, key, f):
        return f in self._load().get(key, {})
    def zadd(self, key, mapping):
        d = self._load()
        if key not in d: d[key] = {}
        for k, v in mapping.items():
            d[key][k] = v
        self._save(d)
    def zrevrange(self, key, start, end):
        z = self._load().get(key, {})
        # Sort keys by score descending
        sorted_keys = sorted(z.keys(), key=lambda k: z[k], reverse=True)
        return sorted_keys
    def zcard(self, key):
        return len(self._load().get(key, {}))
    def zrem(self, key, mem):
        self.hdel(key, mem) # Reusing hdel since zadd uses dict

redis_client = MockRedis()

# Redis key constants
LEADS_KEY = "monitor:leads"              # Hash: lead_id -> JSON
LEADS_SORTED_KEY = "monitor:leads:ts"   # Sorted set: lead_id scored by timestamp
MONITORED_GROUPS_KEY = "monitor:groups" # Set: monitored group JIDs
MONITOR_CONFIG_KEY = "monitor:config"   # Hash: config values

# ─────────────────────────────────────────────
# Internal Helpers
# ─────────────────────────────────────────────

def _save_lead(lead: LeadRecord) -> None:
    """Persist a lead to JSON mock."""
    ts = datetime.fromisoformat(lead.timestamp).timestamp()
    redis_client.hset(LEADS_KEY, lead.lead_id, lead.model_dump_json())
    redis_client.zadd(LEADS_SORTED_KEY, {lead.lead_id: ts})
    logger.info(f"[Monitor] Saved lead: {lead.lead_id} | {lead.phone} | score={lead.score}")


def _get_all_leads() -> list[LeadRecord]:
    """Retrieve all leads from JSON mock."""
    lead_ids = redis_client.zrevrange(LEADS_SORTED_KEY, 0, -1)
    leads = []
    for lid in lead_ids:
        raw = redis_client.hget(LEADS_KEY, lid)
        if raw:
            try:
                leads.append(LeadRecord(**json.loads(raw)))
            except Exception as e:
                logger.warning(f"[Monitor] Could not parse lead {lid}: {e}")
    return leads


def _extract_text_from_webhook(data: dict) -> Optional[str]:
    """Extract the plain text content from an Evolution API webhook payload."""
    message = data.get("message", {}) or {}
    # Try plain text first
    if "conversation" in message:
        return message["conversation"]
    # Extended text
    if "extendedTextMessage" in message:
        return message["extendedTextMessage"].get("text", "")
    # Caption on media
    for media_key in ["imageMessage", "videoMessage", "documentMessage"]:
        if media_key in message:
            caption = message[media_key].get("caption", "")
            if caption:
                return caption
    return None


def _extract_sender_info(data: dict) -> tuple[str, Optional[str]]:
    """
    Extract phone number and display name from webhook payload.
    Returns (phone, name).
    """
    key = data.get("key", {}) or {}
    remote_jid = key.get("remoteJid", "")
    # For group messages, sender is in participant field
    sender_jid = data.get("participant") or key.get("participant") or remote_jid

    # Strip @s.whatsapp.net or @g.us suffix
    phone = sender_jid.replace("@s.whatsapp.net", "").replace("@g.us", "").split("@")[0]

    # Try to get push name
    push_name = data.get("pushName") or data.get("notifyName") or None

    return phone, push_name


# ─────────────────────────────────────────────
# Webhook Receiver
# ─────────────────────────────────────────────

@router.post("/webhook")
async def receive_webhook(request: Request):
    """
    Receives incoming message events from Evolution API.
    Classifies group messages and saves qualifying leads.
    """
    try:
        body = await request.json()
    except Exception:
        return {"status": "ignored", "reason": "invalid json"}

    event = body.get("event", "")

    # Only handle incoming messages
    if event not in ("messages.upsert", "MESSAGES_UPSERT", "message"):
        return {"status": "ignored", "reason": f"unhandled event: {event}"}

    data = body.get("data", {})
    if not data:
        return {"status": "ignored", "reason": "no data"}

    # Handle both single message and array
    messages = data if isinstance(data, list) else [data]

    saved_count = 0
    for msg_data in messages:
        key = msg_data.get("key", {}) or {}
        remote_jid = key.get("remoteJid", "")

        # Only process group messages (JIDs ending in @g.us)
        if not remote_jid.endswith("@g.us"):
            continue

        # Skip messages sent by the bot itself
        if key.get("fromMe", False):
            continue

        # Check if this group is being monitored
        monitored_groups = redis_client.smembers(MONITORED_GROUPS_KEY)
        if monitored_groups and remote_jid not in monitored_groups:
            continue

        # Extract text content
        text = _extract_text_from_webhook(msg_data)
        if not text or not text.strip():
            continue

        # Classify the message
        result = classify_message(text)
        if not result.is_lead:
            continue

        # Extract sender info
        phone, name = _extract_sender_info(msg_data)

        # Get instance name from body
        instance_name = body.get("instance", "bulk-sender-main")

        # Get group name from config cache if available
        group_name_key = f"monitor:group_name:{remote_jid}"
        group_name = redis_client.get(group_name_key) or remote_jid

        lead = LeadRecord(
            lead_id=str(uuid.uuid4()),
            phone=phone,
            name=name,
            message=text[:1000],  # Cap message length
            score=result.score,
            lead_tier=result.lead_tier,
            matched_keywords=result.matched_keywords,
            group_id=remote_jid,
            group_name=group_name,
            timestamp=datetime.now(timezone.utc).isoformat(),
            instance_name=instance_name,
        )

        _save_lead(lead)
        saved_count += 1

    return {"status": "ok", "leads_saved": saved_count}


# ─────────────────────────────────────────────
# Group Endpoints
# ─────────────────────────────────────────────

@router.get("/groups/debug")
async def debug_groups_raw(request: Request, instance_name: str = "bulk-sender-main"):
    """
    Returns the raw Evolution API response for group fetching.
    Use this to diagnose why groups are not loading.
    """
    evo = request.app.state.evolution_api
    raw = await evo.get_groups_raw(instance_name)
    return raw


@router.get("/groups/{group_id:path}/members/debug")
async def debug_group_members(
    group_id: str,
    request: Request,
    instance_name: str = "bulk-sender-main",
):
    """Debug endpoint — returns raw participant data from Baileys."""
    evo = request.app.state.evolution_api
    raw_members = await evo.get_group_members(instance_name, group_id)
    return {"raw_count": len(raw_members), "raw_sample": raw_members[:5], "raw_full": raw_members}


@router.get("/groups")
async def list_groups(
    request: Request,
    instance_name: str = "bulk-sender-main",
    force_refresh: bool = False,
):
    """
    Fetch all WhatsApp groups the connected account belongs to.
    Results are cached in Redis for 5 minutes to avoid the 60-90s Evolution API delay.
    Pass ?force_refresh=true to bypass cache.
    """
    CACHE_KEY = f"monitor:groups_cache:{instance_name}"
    CACHE_TTL = 300  # 5 minutes

    # ── Try cache first (unless force refresh requested)
    if not force_refresh:
        cached = redis_client.get(CACHE_KEY)
        if cached:
            try:
                groups = json.loads(cached)
                logger.info(f"[monitor] Returning {len(groups)} groups from cache")
                return {"groups": groups, "total": len(groups), "cached": True}
            except Exception:
                pass  # Cache corrupt, fall through to fetch

    evo = request.app.state.evolution_api
    try:
        raw_groups = await evo.get_groups(instance_name)
    except Exception as e:
        raise HTTPException(
            status_code=502,
            detail=f"Evolution API error fetching groups: {str(e)}"
        )

    groups = []
    for g in raw_groups:
        gid = g.get("id") or g.get("groupJid") or g.get("remoteJid") or g.get("jid") or ""
        name = g.get("subject") or g.get("name") or g.get("pushName") or gid
        participants = g.get("participants", []) or []
        participant_count = g.get("size") or len(participants)
        desc = g.get("desc") or g.get("description") or None

        if not gid:
            logger.warning(f"[monitor] Skipping group with no ID: {list(g.keys())}")
            continue

        # Cache group name for use by webhook handler
        redis_client.setex(f"monitor:group_name:{gid}", CACHE_TTL * 2, name)

        groups.append(GroupInfo(
            group_id=gid,
            name=name,
            participant_count=participant_count,
            description=desc,
        ).model_dump())

    # ── Store in cache
    redis_client.setex(CACHE_KEY, CACHE_TTL, json.dumps(groups))
    logger.info(f"[monitor] Cached {len(groups)} groups for {CACHE_TTL}s")

    return {"groups": groups, "total": len(groups), "cached": False}


@router.get("/groups/{group_id:path}/members")
async def list_group_members(
    group_id: str,
    request: Request,
    instance_name: str = "bulk-sender-main",
):
    """Fetch all participants of a specific WhatsApp group."""
    evo = request.app.state.evolution_api
    raw_members = await evo.get_group_members(instance_name, group_id)

    members = []
    for m in raw_members:
        # Baileys server now returns {id, phone, admin, idType}
        phone = m.get("phone", "").strip()
        id_type = m.get("idType", "phone")

        # Skip LID entries with no resolvable phone
        if id_type == "lid" or (not phone and "@lid" in (m.get("id") or "")):
            continue

        # Fallback: extract from raw fields if phone not pre-resolved
        if not phone:
            phone_raw = m.get("phoneNumber") or m.get("id") or m.get("jid") or m.get("participant") or ""
            if "@lid" in phone_raw:
                continue
            phone = phone_raw.replace("@s.whatsapp.net", "").replace("@g.us", "").split("@")[0]
            if ":" in phone:
                phone = phone.split(":")[0]

        name = m.get("name") or m.get("pushName") or None
        admin_val = m.get("admin", "") or ""
        is_admin = admin_val in ("admin", "superadmin", True)

        if not phone:
            continue

        members.append(GroupMember(
            phone=phone,
            name=name,
            is_admin=is_admin,
        ).model_dump())

    return {"members": members, "total": len(members)}


@router.get("/groups/{group_id:path}/members/export")
async def export_group_members(
    group_id: str,
    request: Request,
    instance_name: str = "bulk-sender-main",
):
    """
    Export all phone numbers in a WhatsApp group as an Excel (.xlsx) file.
    Ready to be imported back into the Bulk Sender contact list.
    """
    evo = request.app.state.evolution_api
    raw_members = await evo.get_group_members(instance_name, group_id)

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Group Members"

    # ── Header styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["Phone", "Name", "Is Admin", "Group ID"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # ── Data rows
    row_num = 2
    for m in raw_members:
        # Baileys server now returns {id, phone, admin, idType}
        phone = m.get("phone", "").strip()
        id_type = m.get("idType", "phone")

        # Skip LID entries with no resolvable phone
        if id_type == "lid" or (not phone and "@lid" in (m.get("id") or "")):
            continue

        # Fallback: extract from raw fields if phone not pre-resolved
        if not phone:
            phone_raw = m.get("phoneNumber") or m.get("id") or m.get("jid") or m.get("participant") or ""
            if "@lid" in phone_raw:
                continue
            phone = phone_raw.replace("@s.whatsapp.net", "").replace("@g.us", "").split("@")[0]
            if ":" in phone:
                phone = phone.split(":")[0]

        name = m.get("name") or m.get("pushName") or ""
        admin_val = m.get("admin", "") or ""
        is_admin = admin_val in ("admin", "superadmin", True)

        if not phone:
            continue

        ws.cell(row=row_num, column=1, value=phone)
        ws.cell(row=row_num, column=2, value=name)
        ws.cell(row=row_num, column=3, value="Yes" if is_admin else "No")
        ws.cell(row=row_num, column=4, value=group_id)

        # Alternate row coloring
        if row_num % 2 == 0:
            row_fill = PatternFill("solid", fgColor="F0F4FF")
            for col in range(1, 5):
                ws.cell(row=row_num, column=col).fill = row_fill

        row_num += 1

    # Auto-fit columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Stream file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_group = group_id.replace("@g.us", "").replace("@", "")
    filename = f"group_members_{safe_group}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    from fastapi.responses import Response
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────
# Monitor Subscription
# ─────────────────────────────────────────────

@router.post("/subscribe")
async def subscribe_to_groups(
    payload: MonitorSubscription,
    request: Request,
):
    """
    Start monitoring specific groups.
    Registers the webhook with Evolution API and stores group IDs in Redis.
    """
    evo = request.app.state.evolution_api

    # Store monitored group IDs
    if payload.group_ids:
        redis_client.delete(MONITORED_GROUPS_KEY)
        redis_client.sadd(MONITORED_GROUPS_KEY, *payload.group_ids)

    # Attempt to register webhook with Evolution API
    webhook_registered = False
    if payload.webhook_url:
        try:
            await evo.set_webhook(
                instance_name=payload.instance_name,
                webhook_url=payload.webhook_url,
            )
            webhook_registered = True
            redis_client.hset(MONITOR_CONFIG_KEY, "webhook_url", payload.webhook_url)
            redis_client.hset(MONITOR_CONFIG_KEY, "instance_name", payload.instance_name)
        except Exception as e:
            logger.warning(f"[Monitor] Could not register webhook: {e}")

    # Cache group names for display
    for gid in payload.group_ids:
        # We'll cache them when fetched from groups endpoint
        logger.info(f"[Monitor] Now monitoring group: {gid}")

    return {
        "status": "ok",
        "monitored_groups": payload.group_ids,
        "webhook_registered": webhook_registered,
    }


@router.get("/status")
async def get_monitor_status():
    """Return current monitoring configuration and stats."""
    monitored = list(redis_client.smembers(MONITORED_GROUPS_KEY) or [])
    config = redis_client.hgetall(MONITOR_CONFIG_KEY) or {}
    total_leads = redis_client.zcard(LEADS_SORTED_KEY) or 0

    return {
        "monitored_groups": monitored,
        "total_leads": total_leads,
        "webhook_url": config.get("webhook_url"),
        "instance_name": config.get("instance_name"),
        "is_active": len(monitored) > 0,
    }


# ─────────────────────────────────────────────
# Leads Endpoints
# ─────────────────────────────────────────────

@router.get("/leads")
async def get_leads(
    page: int = 1,
    page_size: int = 50,
    group_id: Optional[str] = None,
    tier: Optional[str] = None,
    search: Optional[str] = None,
):
    """
    Retrieve detected leads with optional filtering and pagination.
    Leads are sorted most-recent first.
    """
    leads = _get_all_leads()

    # Apply filters
    if group_id:
        leads = [l for l in leads if l.group_id == group_id]
    if tier:
        leads = [l for l in leads if l.lead_tier == tier]
    if search:
        q = search.lower()
        leads = [
            l for l in leads
            if q in (l.phone or "").lower()
            or q in (l.name or "").lower()
            or q in l.message.lower()
        ]

    total = len(leads)
    start = (page - 1) * page_size
    end = start + page_size
    paginated = leads[start:end]

    return {
        "leads": [l.model_dump() for l in paginated],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, (total + page_size - 1) // page_size),
    }


@router.delete("/leads/{lead_id}")
async def delete_lead(lead_id: str):
    """Remove a single lead from storage."""
    exists = redis_client.hexists(LEADS_KEY, lead_id)
    if not exists:
        raise HTTPException(status_code=404, detail="Lead not found")

    redis_client.hdel(LEADS_KEY, lead_id)
    redis_client.zrem(LEADS_SORTED_KEY, lead_id)
    return {"status": "deleted", "lead_id": lead_id}


@router.delete("/leads")
async def clear_all_leads():
    """Remove all stored leads."""
    redis_client.delete(LEADS_KEY)
    redis_client.delete(LEADS_SORTED_KEY)
    return {"status": "cleared"}


@router.get("/leads/export")
async def export_leads(
    group_id: Optional[str] = None,
    tier: Optional[str] = None,
):
    """
    Export all detected leads as a formatted Excel (.xlsx) file.
    Supports optional filtering by group or lead tier.
    """
    leads = _get_all_leads()

    if group_id:
        leads = [l for l in leads if l.group_id == group_id]
    if tier:
        leads = [l for l in leads if l.lead_tier == tier]

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Real Estate Leads"

    # ── Color palette
    HOT_COLOR = "FF4757"
    WARM_COLOR = "FFA502"
    HEADER_COLOR = "1A1A2E"
    ALT_ROW_COLOR = "FFF8F0"

    # ── Headers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Name", "Phone", "Lead Score", "Tier", "Message",
        "Group", "Matched Keywords", "Timestamp",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 32

    # ── Data rows
    for row_num, lead in enumerate(leads, 2):
        is_hot = lead.lead_tier == "hot"

        ws.cell(row=row_num, column=1, value=lead.name or "Unknown")
        ws.cell(row=row_num, column=2, value=lead.phone)
        ws.cell(row=row_num, column=3, value=lead.score)
        ws.cell(row=row_num, column=4, value=lead.lead_tier.upper())
        ws.cell(row=row_num, column=5, value=lead.message)
        ws.cell(row=row_num, column=6, value=lead.group_name or lead.group_id)
        ws.cell(row=row_num, column=7, value=", ".join(lead.matched_keywords))
        ws.cell(row=row_num, column=8, value=lead.timestamp)

        # Score cell color by tier
        score_cell = ws.cell(row=row_num, column=3)
        score_cell.font = Font(bold=True, color="FFFFFF")
        score_cell.fill = PatternFill("solid", fgColor=HOT_COLOR if is_hot else WARM_COLOR)
        score_cell.alignment = Alignment(horizontal="center")

        # Alternate row background
        if row_num % 2 == 0:
            alt_fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
            for col in [1, 2, 4, 5, 6, 7, 8]:
                ws.cell(row=row_num, column=col).fill = alt_fill

        # Wrap message text
        ws.cell(row=row_num, column=5).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row_num].height = 40

    # ── Column widths
    col_widths = [20, 18, 12, 10, 60, 25, 40, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    # ── Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.append(["Total Leads", len(leads)])
    ws2.append(["Hot Leads (score ≥ 7)", sum(1 for l in leads if l.lead_tier == "hot")])
    ws2.append(["Warm Leads (score 3-6)", sum(1 for l in leads if l.lead_tier == "warm")])
    ws2.append(["Unique Phone Numbers", len({l.phone for l in leads})])
    ws2.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M")])

    # Stream file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"real_estate_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    from fastapi.responses import Response
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
