"""
REKOMND+ — Google Maps Scraper Router
=======================================
Wraps the existing scraper.py logic from gmaps-telegram-bot/
and exposes:
  POST /api/gmaps/search        — start scrape, stream SSE results
  GET  /api/gmaps/sessions      — list all saved scrape sessions
  GET  /api/gmaps/results/{sid} — get all rows for a session
  GET  /api/gmaps/export/{sid}  — download CSV or Excel
  DELETE /api/gmaps/sessions/{sid} — delete session
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import AsyncGenerator

from fastapi import APIRouter, Request
from fastapi.responses import JSONResponse, StreamingResponse, Response

# ── Path magic — import scraper from gmaps-telegram-bot ──────────────────
GMAPS_DIR = Path(__file__).resolve().parents[2] / "gmaps-telegram-bot"
sys.path.insert(0, str(GMAPS_DIR))

try:
    from scraper import GoogleMapsScraper, BusinessInfo  # type: ignore
    SCRAPER_AVAILABLE = True
except ImportError as _e:
    SCRAPER_AVAILABLE = False
    logging.getLogger("gmaps_router").warning("Could not import scraper: %s", _e)

logger = logging.getLogger("gmaps_router")
router = APIRouter()

# ---------------------------------------------------------------------------
# SQLite Database
# ---------------------------------------------------------------------------

DB_PATH = Path(__file__).resolve().parents[1] / "data" / "gmaps_results.db"
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

def _get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn

def _init_db() -> None:
    with _get_conn() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS sessions (
                id          TEXT PRIMARY KEY,
                query       TEXT NOT NULL,
                created_at  TEXT NOT NULL,
                row_count   INTEGER DEFAULT 0
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS results (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id          TEXT NOT NULL,
                name                TEXT,
                rating              REAL,
                review_count        INTEGER,
                category            TEXT,
                price_level         TEXT,
                address             TEXT,
                phone               TEXT,
                website             TEXT,
                maps_url            TEXT,
                reviews_scraped     INTEGER,
                scraped_at          TEXT,
                FOREIGN KEY (session_id) REFERENCES sessions(id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS gmaps_reviews (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                result_id   INTEGER NOT NULL,
                author      TEXT,
                rating      INTEGER,
                text        TEXT,
                date        TEXT,
                likes       INTEGER,
                FOREIGN KEY (result_id) REFERENCES results(id) ON DELETE CASCADE
            )
        """)
        conn.commit()

_init_db()

def _save_business(session_id: str, info: "BusinessInfo") -> int:
    with _get_conn() as conn:
        cur = conn.execute("""
            INSERT INTO results
            (session_id, name, rating, review_count, category, price_level,
             address, phone, website, maps_url, reviews_scraped, scraped_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            session_id,
            info.name,
            info.rating,
            info.review_count,
            info.category,
            info.price_level,
            info.address,
            info.phone,
            info.website,
            info.maps_url,
            len(info.reviews),
            datetime.utcnow().isoformat(),
        ))
        row_id = cur.lastrowid

        if info.reviews:
            review_data = [(row_id, r.author, r.rating, r.text, r.date, r.likes) for r in info.reviews]
            conn.executemany("""
                INSERT INTO gmaps_reviews (result_id, author, rating, text, date, likes)
                VALUES (?, ?, ?, ?, ?, ?)
            """, review_data)

        conn.execute(
            "UPDATE sessions SET row_count = row_count + 1 WHERE id = ?",
            (session_id,)
        )
        conn.commit()
        return row_id

# ---------------------------------------------------------------------------
# SSE helper
# ---------------------------------------------------------------------------

def _sse(event: str, data: dict) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"

# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/search")
async def start_search(request: Request):
    """Start a scraping job — returns Server-Sent Events stream."""
    body = await request.json()
    query: str = body.get("query", "").strip()
    max_businesses: int = int(body.get("max_businesses", 50))
    max_reviews: int    = int(body.get("max_reviews", 20))

    if not query:
        return JSONResponse({"error": "query is required"}, status_code=400)

    if not SCRAPER_AVAILABLE:
        return JSONResponse({"error": "Scraper module not available. Make sure playwright is installed."}, status_code=503)

    # Create session record
    session_id = f"{int(time.time())}_{query[:20].replace(' ', '_')}"
    with _get_conn() as conn:
        conn.execute(
            "INSERT INTO sessions (id, query, created_at) VALUES (?, ?, ?)",
            (session_id, query, datetime.utcnow().isoformat()),
        )
        conn.commit()

    async def generate() -> AsyncGenerator[str, None]:
        yield _sse("session_start", {"session_id": session_id, "query": query})

        count = 0
        try:
            scraper = GoogleMapsScraper(max_reviews=max_reviews, headless=True)
            async for info in scraper.search_multiple(query, max_businesses=max_businesses):
                # Stop if user clicked Stop / closed page
                if await request.is_disconnected():
                    logger.info("Client disconnected — stopping scrape early.")
                    break

                if info.error:
                    if count == 0:
                        yield _sse("error", {"message": info.error})
                        return
                    continue

                count += 1
                row_id = _save_business(session_id, info)

                row = {
                    "id": row_id,
                    "name": info.name,
                    "rating": info.rating,
                    "review_count": info.review_count,
                    "category": info.category,
                    "price_level": info.price_level,
                    "address": info.address,
                    "phone": info.phone,
                    "website": info.website,
                    "maps_url": info.maps_url,
                    "reviews_scraped": len(info.reviews),
                }
                yield _sse("business", row)

                # Yield control so SSE flushes
                await asyncio.sleep(0)

        except Exception as exc:
            logger.exception("Scrape error: %s", exc)
            yield _sse("error", {"message": str(exc)})
        finally:
            yield _sse("done", {"count": count, "session_id": session_id})

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@router.get("/sessions")
def list_sessions():
    with _get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM sessions ORDER BY created_at DESC LIMIT 100"
        ).fetchall()
    return [dict(r) for r in rows]


@router.get("/results/{session_id}")
def get_results(session_id: str):
    with _get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM results WHERE session_id = ? ORDER BY id",
            (session_id,),
        ).fetchall()
    return [dict(r) for r in rows]


@router.get("/reviews/{result_id}")
def get_reviews(result_id: int):
    with _get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM gmaps_reviews WHERE result_id = ? ORDER BY id",
            (result_id,),
        ).fetchall()
    return [dict(r) for r in rows]


@router.get("/export/{session_id}")
def export_results(session_id: str, format: str = "csv"):
    with _get_conn() as conn:
        session = conn.execute(
            "SELECT * FROM sessions WHERE id = ?", (session_id,)
        ).fetchone()
        rows = conn.execute(
            "SELECT * FROM results WHERE session_id = ? ORDER BY id",
            (session_id,),
        ).fetchall()

    if not session:
        return JSONResponse({"error": "Session not found"}, status_code=404)

    query = dict(session)["query"]
    safe_query = "".join(c if c.isalnum() or c in " _-" else "" for c in query)[:40]
    filename_base = f"rekomnd_gmaps_{safe_query}_{session_id[:12]}"

    headers_list = [
        "Name", "Rating", "Reviews", "Category", "Price Level",
        "Address", "Phone", "Website", "Maps URL", "Reviews Scraped", "Scraped At"
    ]
    keys = [
        "name", "rating", "review_count", "category", "price_level",
        "address", "phone", "website", "maps_url", "reviews_scraped", "scraped_at"
    ]

    data_rows = [dict(r) for r in rows]

    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "GMaps Results"

            # Header styling
            header_fill = PatternFill("solid", fgColor="1a1f3a")
            header_font = Font(bold=True, color="F59E0B")
            for col_idx, h in enumerate(headers_list, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row_data in enumerate(data_rows, 2):
                for col_idx, key in enumerate(keys, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ""))

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
            )
        except ImportError:
            pass  # Fall through to CSV

    # CSV
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers_list)
    for row_data in data_rows:
        writer.writerow([row_data.get(k, "") for k in keys])

    content = buf.getvalue().encode("utf-8-sig")  # BOM for Excel
    return Response(
        content=content,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
    )


@router.delete("/sessions/{session_id}")
def delete_session(session_id: str):
    with _get_conn() as conn:
        # Get result IDs for this session to delete reviews
        rows = conn.execute("SELECT id FROM results WHERE session_id = ?", (session_id,)).fetchall()
        result_ids = [r["id"] for r in rows]
        if result_ids:
            conn.execute(f"DELETE FROM gmaps_reviews WHERE result_id IN ({','.join('?'*len(result_ids))})", result_ids)
            
        conn.execute("DELETE FROM results WHERE session_id = ?", (session_id,))
        conn.execute("DELETE FROM sessions WHERE id = ?", (session_id,))
        conn.commit()
    return {"deleted": session_id}
